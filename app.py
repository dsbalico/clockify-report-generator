# Import the necessary libraries
from openpyxl import load_workbook
from datetime import datetime
from colorama import Fore, Style
import argparse

# Import the functions from the other files
from utils.logger import initialize_logger
from configuration import Config
from utils.api_client import ClockifyAPIClient
from utils.data_processing import (
    get_period_covered,
    transform_data,
    get_week_number
)

logger = initialize_logger()
config = Config()

def write_to_excel(transformed_data, period_covered, user_name):
    file_path = 'template.xlsx'
    workbook = load_workbook(file_path)

    sheet = workbook.active

    sheet["B1"] = user_name
    sheet["B2"] = config.POSITION
    sheet["B3"] = period_covered

    start_row = 6
    start_col = 1  

    current_row = start_row
    last_date = None  # To keep track of the previous date

    for entry in transformed_data:
        date = entry['date']
        description = entry['description']
        startTime = entry['startTime']
        endTime = entry['endTime']
        duration = entry['duration']

        if date != last_date:
            if last_date is not None:
                start_row += 8
                current_row = start_row
            
            last_date = date
        else:
            last_date = date
            date = ''  # Don't repeat the date if it's the same as the previous one

        sheet.cell(row=current_row, column=start_col, value=date)               # A column
        sheet.cell(row=current_row, column=start_col + 1, value=description)    # B column
        sheet.cell(row=current_row, column=start_col + 2, value=startTime)      # C column
        sheet.cell(row=current_row, column=start_col + 3, value=endTime)        # D column
        sheet.cell(row=current_row, column=start_col + 4, value=duration)       # E column

        current_row += 1

    filename = generate_filename(transformed_data, user_name)

    workbook.save(filename)

    logger.info(f"File saved as: {Fore.GREEN}{filename}{Style.RESET_ALL}")

def generate_filename(transformed_data, user_name):
    date_obj = datetime.strptime(transformed_data[0]['date'], '%m/%d/%Y')
    formatted_date = date_obj.strftime('%m%Y')
    week_number = get_week_number(date_obj)

    names = user_name.split(' ')

    if len(names) == 2:
        first_name = names[0]
        last_name = names[1]
        middle_initial = ''
    elif len(names) == 3:
        first_name = names[0]
        middle_initial = names[1][0]
        last_name = names[2]
    elif len(names) == 4:
        first_name = names[0] + ' ' + names[1]
        middle_initial = names[2][0]
        last_name = names[3]
    elif len(names) == 5:
        first_name = names[0] + ' ' + names[1] + ' ' + names[2]
        middle_initial = names[3][0]
        last_name = names[4]
    else:
        first_name = '---'
        middle_initial = '---'
        last_name = '---'

    output_file_path = f'output/Accomplishment Report [{formatted_date} Week#{week_number}] - {last_name}, {first_name}, {middle_initial}..xlsx'

    return output_file_path

def main():
    parser = argparse.ArgumentParser(description='Process some command-line arguments.')
    parser.add_argument('--week', type=str, help='The week number or specific date to generate the report for.')
    parser.add_argument('--start_date', type=str, help='The start date for the range (YYYY-MM-DD).')
    parser.add_argument('--end_date', type=str, help='The end date for the range (YYYY-MM-DD).')
    args = parser.parse_args()

    logger.info(f"============== {Fore.GREEN}[Starting Clockify Report Generator]{Style.RESET_ALL} ==============")

    api_client = ClockifyAPIClient(config.API_KEY, config.BASE_URL, config.WORKSPACE_ID, config.USER_ID)

    if args.start_date and args.end_date:
        logger.info(f"Generating report for date range: {Fore.CYAN}{args.start_date} to {args.end_date}{Style.RESET_ALL}")
        try:
            start_date = datetime.strptime(args.start_date, '%Y-%m-%d')
            end_date = datetime.strptime(args.end_date, '%Y-%m-%d')
        except ValueError:
            logger.error("Invalid date format. Please use YYYY-MM-DD.")
            return
        time_entries = api_client.get_time_entries_by_date_range(start_date, end_date)
    elif args.week:
        try:
            input_date = datetime.strptime(args.week, '%Y-%m-%d')
            logger.info(f"Generating report for date: {Fore.CYAN}{input_date.strftime('%Y-%m-%d')}{Style.RESET_ALL}")
            time_entries = api_client.get_time_entries_by_date(input_date)
        except ValueError:
            logger.info(f"Generating WAR for Week: {Fore.CYAN}{args.week}{Style.RESET_ALL}")
            time_entries = api_client.get_time_entries(args.week)
    else:
        logger.error("Please provide either a week option or a date range.")
        return

    # Process and write data as before
    logger.info(f"Getting User Info...")
    user = api_client.get_user_info()
    logger.info(f"User ID: {Fore.CYAN}{user['id']}{Style.RESET_ALL} | User Name: {Fore.CYAN}{user['name']}{Style.RESET_ALL} | User Email: {Fore.CYAN}{user['email']}{Style.RESET_ALL}")

    logger.info(f"Getting Workspaces...")
    workspaces = api_client.get_workspaces()
    for workspace in workspaces:
        logger.info(f"Workspace ID: {Fore.CYAN}{workspace['id']}{Style.RESET_ALL} | Workspace Name: {Fore.CYAN}{workspace['name']}{Style.RESET_ALL}")

    logger.info(f"Fetching Time Entries...")

    logger.info(f"{Fore.MAGENTA}Processing Data...{Style.RESET_ALL}")

    period_covered = get_period_covered(time_entries)
    transformed_data = transform_data(time_entries)
    
    logger.info(f"{Fore.MAGENTA}Writing to Excel...{Style.RESET_ALL}")

    write_to_excel(transformed_data, period_covered, user['name'])

    logger.info(f"============================== {Fore.GREEN}[DONE]{Style.RESET_ALL} =============================")

if __name__ == '__main__':
    main()